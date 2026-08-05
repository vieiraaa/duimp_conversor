from pathlib import Path
from dataclasses import asdict

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo

import pandas as pd


class ExcelExporter:

    def __init__(self, resumo, itens):

        self.resumo = resumo
        self.itens = itens

    def exportar(self, arquivo):

        Path(arquivo).parent.mkdir(
            parents=True,
            exist_ok=True
        )

        resumo_df = pd.DataFrame(
            list(asdict(self.resumo).items()),
            columns=["Campo", "Valor"]
        )

        itens_df = pd.DataFrame(
            [asdict(item) for item in self.itens]
        )

        with pd.ExcelWriter(
            arquivo,
            engine="openpyxl"
        ) as writer:

            resumo_df.to_excel(
                writer,
                sheet_name="Resumo",
                index=False
            )

            itens_df.to_excel(
                writer,
                sheet_name="Itens",
                index=False
            )

        self.formatar(arquivo)

    def formatar(self, arquivo):

        wb = load_workbook(arquivo)

        azul = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        fonte = Font(
            bold=True,
            color="FFFFFF",
            size=11
        )

        borda = Border(

            left=Side(style="thin"),

            right=Side(style="thin"),

            top=Side(style="thin"),

            bottom=Side(style="thin")

        )

        for ws in wb.worksheets:

            # Cabeçalho

            for cell in ws[1]:

                cell.fill = azul
                cell.font = fonte
                cell.border = borda
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # Linhas

            for row in ws.iter_rows(min_row=2):

                for cell in row:

                    cell.border = borda

            # Ajuste automático

            for coluna in ws.columns:

                largura = 0

                letra = get_column_letter(
                    coluna[0].column
                )

                for cell in coluna:

                    try:

                        largura = max(
                            largura,
                            len(str(cell.value))
                        )

                    except:

                        pass

                ws.column_dimensions[letra].width = min(
                    largura + 4,
                    60
                )

            ws.freeze_panes = "A2"

            ws.auto_filter.ref = ws.dimensions

        self._formatar_itens(wb["Itens"])

        wb.save(arquivo)

    def _formatar_itens(self, ws):

        ultima_linha = ws.max_row
        ultima_coluna = ws.max_column

        referencia = (
            f"A1:"
            f"{get_column_letter(ultima_coluna)}"
            f"{ultima_linha}"
        )

        tabela = Table(

            displayName="Itens",

            ref=referencia

        )

        estilo = TableStyleInfo(

            name="TableStyleMedium2",

            showFirstColumn=False,

            showLastColumn=False,

            showRowStripes=True,

            showColumnStripes=False

        )

        tabela.tableStyleInfo = estilo

        ws.add_table(tabela)

        # Quantidade

        for cell in ws["G"][1:]:

            cell.number_format = "0.000"

        # Peso Líquido

        for cell in ws["H"][1:]:

            cell.number_format = "0.000"

        # Peso Bruto

        for cell in ws["I"][1:]:

            cell.number_format = "0.000"

        # Valor Unitário

        for cell in ws["J"][1:]:

            cell.number_format = 'R$ #,##0.00'

        # Valor Total

        for cell in ws["K"][1:]:

            cell.number_format = 'R$ #,##0.00'

        linha_total = ultima_linha + 2

        ws[f"F{linha_total}"] = "TOTAL"

        ws[f"G{linha_total}"] = (
            f"=SUM(G2:G{ultima_linha})"
        )

        ws[f"H{linha_total}"] = (
            f"=SUM(H2:H{ultima_linha})"
        )

        ws[f"I{linha_total}"] = (
            f"=SUM(I2:I{ultima_linha})"
        )

        ws[f"K{linha_total}"] = (
            f"=SUM(K2:K{ultima_linha})"
        )

        for coluna in ["F", "G", "H", "I", "K"]:

            ws[f"{coluna}{linha_total}"].font = Font(
                bold=True
            )